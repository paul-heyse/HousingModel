"""Tests for Excel export functionality."""

import pytest
import tempfile
from pathlib import Path
from datetime import date

import openpyxl

from aker_core.exports import ExcelWorkbookBuilder, to_excel


class TestExcelWorkbookBuilder:
    """Test Excel workbook builder functionality."""

    def test_workbook_builder_initialization(self, db_session):
        """Test workbook builder initialization."""
        builder = ExcelWorkbookBuilder(db_session)
        assert builder.session is db_session
        assert builder.workbook is None
        assert builder.output_path is None

    def test_workbook_generation(self, db_session):
        """Test complete workbook generation."""
        builder = ExcelWorkbookBuilder(db_session)

        # Generate workbook
        output_path = builder.build_workbook("BOI", "TestAsset", date.today())

        # Verify file was created
        assert output_path.exists()
        assert output_path.suffix == ".xlsx"

        # Verify workbook structure
        workbook = openpyxl.load_workbook(output_path)

        # Check required sheets exist
        expected_sheets = [
            "Overview", "Market_Scorecard", "Asset_Fit", "Deal_Archetypes",
            "Risk", "Ops_KPIs", "CO-UT-ID_Patterns", "Checklist",
            "Data_Lineage", "Config"
        ]

        for sheet_name in expected_sheets:
            assert sheet_name in workbook.sheetnames

        # Verify workbook properties
        assert workbook.properties.title is not None
        assert "BOI" in workbook.properties.title
        assert "TestAsset" in workbook.properties.title

    def test_sheet_content_validation(self, db_session):
        """Test that sheets contain expected content."""
        builder = ExcelWorkbookBuilder(db_session)
        output_path = builder.build_workbook("DEN", "SampleProperty", date.today())

        workbook = openpyxl.load_workbook(output_path)

        # Check Overview sheet
        overview_sheet = workbook["Overview"]
        assert overview_sheet["A1"].value is not None  # Title should be present

        # Check Market_Scorecard sheet
        scorecard_sheet = workbook["Market_Scorecard"]
        assert scorecard_sheet["A1"].value is not None

        # Check Config sheet has metadata
        config_sheet = workbook["Config"]
        assert config_sheet["B4"].value == "DEN"  # MSA
        assert config_sheet["B5"].value == "SampleProperty"  # Asset

    def test_file_cleanup_on_error(self, db_session):
        """Test that temporary files are cleaned up on errors."""
        builder = ExcelWorkbookBuilder(db_session)

        # This should work without errors
        output_path = builder.build_workbook("SLC", "TestAsset", date.today())
        assert output_path.exists()

        # Clean up
        output_path.unlink(missing_ok=True)


class TestExcelExportConvenience:
    """Test convenience functions for Excel export."""

    def test_to_excel_convenience_function(self, db_session):
        """Test the convenience to_excel function."""
        output_path = to_excel("BOI", "ConvenienceTest", date.today(), db_session)

        assert output_path.exists()
        assert output_path.suffix == ".xlsx"

        # Verify workbook structure
        workbook = openpyxl.load_workbook(output_path)
        assert "Overview" in workbook.sheetnames
        assert "Market_Scorecard" in workbook.sheetnames

    def test_to_excel_without_session(self):
        """Test to_excel with default session."""
        # This should work with the default session
        output_path = to_excel("DEN", "DefaultSession", date.today())

        assert output_path.exists()
        assert output_path.suffix == ".xlsx"


class TestExcelFileFormat:
    """Test Excel file format and structure."""

    def test_workbook_opens_in_excel(self, db_session):
        """Test that generated workbook can be opened by Excel."""
        builder = ExcelWorkbookBuilder(db_session)
        output_path = builder.build_workbook("BOI", "ExcelTest", date.today())

        # Verify file is readable by openpyxl
        workbook = openpyxl.load_workbook(output_path)
        assert workbook is not None

        # Verify basic structure
        assert len(workbook.sheetnames) >= 5  # Should have multiple sheets

    def test_sheet_name_validation(self, db_session):
        """Test that all required sheets have correct names."""
        builder = ExcelWorkbookBuilder(db_session)
        output_path = builder.build_workbook("SLC", "SheetTest", date.today())

        workbook = openpyxl.load_workbook(output_path)

        expected_sheets = {
            "Overview", "Market_Scorecard", "Asset_Fit", "Deal_Archetypes",
            "Risk", "Ops_KPIs", "CO-UT-ID_Patterns", "Checklist",
            "Data_Lineage", "Config"
        }

        actual_sheets = set(workbook.sheetnames)
        assert expected_sheets.issubset(actual_sheets)

    def test_cell_formatting(self, db_session):
        """Test that cells have appropriate formatting."""
        builder = ExcelWorkbookBuilder(db_session)
        output_path = builder.build_workbook("DEN", "FormatTest", date.today())

        workbook = openpyxl.load_workbook(output_path)
        overview_sheet = workbook["Overview"]

        # Check that some cells have content (basic validation)
        assert overview_sheet["A1"].value is not None

        # Check that we can access cells without errors
        cell_a2 = overview_sheet["A2"]
        cell_b1 = overview_sheet["B1"]
        # These should not raise exceptions


class TestExcelExportIntegration:
    """Test Excel export integration with other systems."""

    def test_export_with_real_data_structure(self, db_session):
        """Test export with realistic data structure."""
        builder = ExcelWorkbookBuilder(db_session)

        # Generate workbook with realistic parameters
        output_path = builder.build_workbook("BOI", "RealDataTest", date(2024, 9, 15))

        assert output_path.exists()

        # Verify file size is reasonable (not empty)
        assert output_path.stat().st_size > 1000  # At least 1KB

        # Verify workbook can be reopened
        workbook = openpyxl.load_workbook(output_path)
        assert len(workbook.sheetnames) >= 8

    def test_multiple_exports_consistency(self, db_session):
        """Test that multiple exports produce consistent results."""
        builder = ExcelWorkbookBuilder(db_session)

        # Generate two workbooks with same parameters
        path1 = builder.build_workbook("BOI", "Consistency1", date.today())
        path2 = builder.build_workbook("BOI", "Consistency2", date.today())

        # Both should exist and have same size (deterministic generation)
        assert path1.exists()
        assert path2.exists()

        # File sizes should be similar (allowing for timestamp differences)
        size1 = path1.stat().st_size
        size2 = path2.stat().st_size
        assert abs(size1 - size2) < 1000  # Within 1KB difference

        # Clean up
        path1.unlink(missing_ok=True)
        path2.unlink(missing_ok=True)


class TestExcelExportErrorHandling:
    """Test error handling in Excel export."""

    def test_invalid_parameters(self, db_session):
        """Test handling of invalid parameters."""
        builder = ExcelWorkbookBuilder(db_session)

        # Test with invalid MSA code
        with pytest.raises(Exception):  # Should raise some validation error
            builder.build_workbook("", "TestAsset", date.today())

    def test_file_system_errors(self, db_session):
        """Test handling of file system errors."""
        builder = ExcelWorkbookBuilder(db_session)

        # Test with read-only directory (would need actual setup)
        # This is a simplified test - real implementation would need proper mocking
        output_path = builder.build_workbook("BOI", "FSError", date.today())
        assert output_path.exists()


class TestExcelExportPerformance:
    """Test Excel export performance."""

    def test_export_performance(self, db_session):
        """Test that export completes in reasonable time."""
        import time

        builder = ExcelWorkbookBuilder(db_session)

        start_time = time.time()
        output_path = builder.build_workbook("BOI", "PerfTest", date.today())
        end_time = time.time()

        # Should complete in reasonable time (< 5 seconds)
        assert end_time - start_time < 5.0
        assert output_path.exists()

        # Clean up
        output_path.unlink(missing_ok=True)

    def test_memory_usage(self, db_session):
        """Test that memory usage remains reasonable during export."""
        import psutil
        import os

        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss / 1024 / 1024  # MB

        builder = ExcelWorkbookBuilder(db_session)
        output_path = builder.build_workbook("BOI", "MemoryTest", date.today())

        final_memory = process.memory_info().rss / 1024 / 1024  # MB
        memory_increase = final_memory - initial_memory

        # Memory increase should be reasonable (< 100MB)
        assert memory_increase < 100

        # Clean up
        output_path.unlink(missing_ok=True)

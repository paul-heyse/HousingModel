from __future__ import annotations

from abc import ABC, abstractmethod
from datetime import date
from typing import Dict, List, Any, Optional
from sqlalchemy.orm import Session

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class SheetBuilder(ABC):
    """Abstract base class for Excel sheet builders."""

    def __init__(self, session: Session, workbook: Workbook):
        self.session = session
        self.workbook = workbook
        self.sheet = None

    @abstractmethod
    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the specific sheet."""
        pass

    def _create_sheet(self, name: str):
        """Create a new worksheet."""
        self.sheet = self.workbook.create_sheet(name)

    def _write_cell(self, row: int, col: int, value: Any, style: str = "normal"):
        """Write a cell with optional styling."""
        cell = self.sheet.cell(row=row, column=col, value=value)

        # Apply basic styling
        if style == "header":
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        elif style == "subheader":
            cell.font = Font(bold=True, size=10)
        elif style == "metric":
            cell.font = Font(bold=True, color="0066CC")

        return cell

    def _write_data_range(self, start_row: int, start_col: int, data: List[List[Any]]):
        """Write a range of data cells."""
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_value in enumerate(row_data):
                self._write_cell(
                    start_row + row_idx,
                    start_col + col_idx,
                    cell_value
                )

    def _add_chart(self, chart_type: str, title: str, data_range: str, position: tuple):
        """Add a chart to the sheet."""
        # Simplified chart creation - would need full implementation
        pass


class OverviewSheet(SheetBuilder):
    """Builder for the Overview sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Overview sheet with executive summary."""
        self._create_sheet("Overview")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, f"Aker Property Model - {msa} - {asset}", "header")
        current_row += 2

        # Executive Summary
        self._write_cell(current_row, 1, "Executive Summary", "subheader")
        current_row += 1

        # Key metrics (would be populated from actual data)
        metrics = [
            ["Market Score", "4.2", "/5"],
            ["Asset Fit Score", "87", "/100"],
            ["Risk Multiplier", "0.95", ""],
            ["Estimated IRR", "12.5%", ""],
            ["Payback Period", "3.2", "years"]
        ]

        for metric in metrics:
            self._write_cell(current_row, 1, metric[0])
            self._write_cell(current_row, 2, metric[1], "metric")
            self._write_cell(current_row, 3, metric[2])
            current_row += 1

        current_row += 1

        # Recommendations
        self._write_cell(current_row, 1, "Key Recommendations", "subheader")
        current_row += 1

        recommendations = [
            "• Proceed with value-add strategy focusing on unit upgrades",
            "• Implement sustainability features for long-term value",
            "• Monitor local employment trends for rent optimization",
            "• Consider EV charging infrastructure for competitive advantage"
        ]

        for rec in recommendations:
            self._write_cell(current_row, 1, rec)
            current_row += 1


class MarketScorecardSheet(SheetBuilder):
    """Builder for the Market_Scorecard sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Market_Scorecard sheet with pillar analysis."""
        self._create_sheet("Market_Scorecard")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, f"Market Scorecard - {msa}", "header")
        current_row += 2

        # Pillar scores
        self._write_cell(current_row, 1, "Pillar Analysis", "subheader")
        current_row += 1

        # Pillar headers
        headers = ["Pillar", "Score (0-5)", "Normalized (0-100)", "Weight", "Weighted Score"]
        for col, header in enumerate(headers):
            self._write_cell(current_row, col + 1, header, "header")
        current_row += 1

        # Sample pillar data (would be populated from actual market scoring)
        pillars = [
            ["Supply Constraint", "3.8", "76", "30%", "2.28"],
            ["Innovation Jobs", "4.2", "84", "30%", "2.52"],
            ["Urban Convenience", "3.5", "70", "20%", "1.40"],
            ["Outdoor Access", "4.1", "82", "20%", "1.64"],
            ["", "Market Score:", "4.2", "", "7.84/10"]
        ]

        for pillar in pillars:
            for col, value in enumerate(pillar):
                if col == 0 and value == "":  # Total row
                    self._write_cell(current_row, col + 1, value, "subheader")
                elif col == 1 and value != "Market Score:":
                    self._write_cell(current_row, col + 1, value, "metric")
                else:
                    self._write_cell(current_row, col + 1, value)
            current_row += 1


class AssetFitSheet(SheetBuilder):
    """Builder for the Asset_Fit sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Asset_Fit sheet with asset evaluation."""
        self._create_sheet("Asset_Fit")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, f"Asset Fit Analysis - {asset}", "header")
        current_row += 2

        # Asset overview
        self._write_cell(current_row, 1, "Asset Overview", "subheader")
        current_row += 1

        overview_data = [
            ["Property Type", "Garden Style"],
            ["Units", "150"],
            ["Year Built", "2010"],
            ["Current Occupancy", "92%"],
            ["Market Position", "Value-Add Opportunity"]
        ]

        for item in overview_data:
            self._write_cell(current_row, 1, item[0])
            self._write_cell(current_row, 2, item[1])
            current_row += 1

        current_row += 1

        # Fit scores
        self._write_cell(current_row, 1, "Fit Scores", "subheader")
        current_row += 1

        fit_data = [
            ["Product Type Compatibility", "85", "/100"],
            ["Vintage Suitability", "78", "/100"],
            ["Unit Mix Optimization", "82", "/100"],
            ["Amenity Enhancement", "88", "/100"],
            ["Overall Asset Fit", "83", "/100"]
        ]

        for fit in fit_data:
            self._write_cell(current_row, 1, fit[0])
            self._write_cell(current_row, 2, fit[1], "metric")
            self._write_cell(current_row, 3, fit[2])
            current_row += 1


class DealArchetypesSheet(SheetBuilder):
    """Builder for the Deal_Archetypes sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Deal_Archetypes sheet with deal analysis."""
        self._create_sheet("Deal_Archetypes")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, "Deal Archetype Analysis", "header")
        current_row += 2

        # Archetype comparison
        self._write_cell(current_row, 1, "Recommended Archetypes", "subheader")
        current_row += 1

        headers = ["Archetype", "Cost/Unit", "Expected Lift", "Payback (years)", "ROI Score"]
        for col, header in enumerate(headers):
            self._write_cell(current_row, col + 1, header, "header")
        current_row += 1

        archetypes = [
            ["Value-Add Interiors", "$15,000", "+$120/mo", "3.2", "8.5/10"],
            ["Amenity Enhancement", "$25,000", "+$180/mo", "2.8", "9.2/10"],
            ["Sustainability Upgrade", "$35,000", "+$150/mo", "4.1", "7.8/10"]
        ]

        for archetype in archetypes:
            for col, value in enumerate(archetype):
                if col == 4:  # ROI Score
                    self._write_cell(current_row, col + 1, value, "metric")
                else:
                    self._write_cell(current_row, col + 1, value)
            current_row += 1


class RiskSheet(SheetBuilder):
    """Builder for the Risk sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Risk sheet with risk assessment."""
        self._create_sheet("Risk")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, "Risk Assessment", "header")
        current_row += 2

        # Risk categories
        self._write_cell(current_row, 1, "Risk Categories", "subheader")
        current_row += 1

        risk_data = [
            ["Market Risk", "Medium", "2.5/5", "Monitor employment trends"],
            ["Construction Risk", "Low", "1.8/5", "Standard contractor oversight"],
            ["Regulatory Risk", "Medium", "2.2/5", "Track permit timelines"],
            ["Environmental Risk", "High", "3.8/5", "Enhanced insurance required"],
            ["Overall Risk Score", "Medium-High", "2.6/5", "Proceed with enhanced monitoring"]
        ]

        for risk in risk_data:
            for col, value in enumerate(risk):
                if col == 2:  # Risk score
                    self._write_cell(current_row, col + 1, value, "metric")
                else:
                    self._write_cell(current_row, col + 1, value)
            current_row += 1


class OpsKPIsSheet(SheetBuilder):
    """Builder for the Ops_KPIs sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Ops_KPIs sheet with operational metrics."""
        self._create_sheet("Ops_KPIs")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, "Operational KPIs", "header")
        current_row += 2

        # KPI categories
        self._write_cell(current_row, 1, "Performance Metrics", "subheader")
        current_row += 1

        kpi_data = [
            ["Reputation Index", "78", "/100", "Above average"],
            ["Average Lease Term", "14", "months", "Stable"],
            ["Renewal Rate", "68%", "", "Improving"],
            ["Maintenance Response", "2.3", "days", "Industry standard"],
            ["Operating Margin", "32%", "", "Target range"]
        ]

        for kpi in kpi_data:
            for col, value in enumerate(kpi):
                if col == 1:  # KPI value
                    self._write_cell(current_row, col + 1, value, "metric")
                else:
                    self._write_cell(current_row, col + 1, value)
            current_row += 1


class COUTIDPatternsSheet(SheetBuilder):
    """Builder for the CO-UT-ID_Patterns sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the CO-UT-ID_Patterns sheet with state-specific data."""
        self._create_sheet("CO-UT-ID_Patterns")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, "State-Specific Operational Patterns", "header")
        current_row += 2

        # State characteristics
        self._write_cell(current_row, 1, "Regional Characteristics", "subheader")
        current_row += 1

        state_data = [
            ["Colorado", "Tech/Health anchors", "Hail/Wildfire focus", "Annual tax cadence"],
            ["Utah", "Higher-ed concentration", "Water rights complexity", "Biennial assessment"],
            ["Idaho", "Migration-driven growth", "Forest interface risks", "5-year cycles"]
        ]

        for state in state_data:
            for col, value in enumerate(state):
                self._write_cell(current_row, col + 1, value)
            current_row += 1


class ChecklistSheet(SheetBuilder):
    """Builder for the Checklist sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Checklist sheet with due diligence tracking."""
        self._create_sheet("Checklist")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, "Due Diligence Checklist", "header")
        current_row += 2

        # Checklist categories
        categories = ["Market", "Site", "Building", "Financial/Ops"]

        for category in categories:
            self._write_cell(current_row, 1, f"{category} Analysis", "subheader")
            current_row += 1

            # Sample checklist items
            items = [
                f"• {category.lower()} analysis review", "Complete", "100%",
                f"• {category.lower()} risk assessment", "In Progress", "75%",
                f"• {category.lower()} compliance check", "Pending", "0%"
            ]

            for item in items:
                parts = item.split(", ")
                self._write_cell(current_row, 1, parts[0])
                self._write_cell(current_row, 2, parts[1])
                self._write_cell(current_row, 3, parts[2])
                current_row += 1

            current_row += 1


class DataLineageSheet(SheetBuilder):
    """Builder for the Data_Lineage sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Data_Lineage sheet with provenance tracking."""
        self._create_sheet("Data_Lineage")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, "Data Provenance & Lineage", "header")
        current_row += 2

        # Data sources
        self._write_cell(current_row, 1, "Data Sources", "subheader")
        current_row += 1

        sources = [
            ["Census ACS", "Population & demographics", "Monthly", "2024-08-15"],
            ["BLS CES", "Employment data", "Monthly", "2024-08-15"],
            ["HUD", "Housing vacancy", "Quarterly", "2024-07-01"],
            ["OSM", "POI & network data", "Weekly", "2024-09-20"]
        ]

        headers = ["Source", "Data Type", "Update Frequency", "Last Update"]
        for col, header in enumerate(headers):
            self._write_cell(current_row, col + 1, header, "header")
        current_row += 1

        for source in sources:
            for col, value in enumerate(source):
                self._write_cell(current_row, col + 1, value)
            current_row += 1


class ConfigSheet(SheetBuilder):
    """Builder for the Config sheet."""

    def build_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Config sheet with system metadata."""
        self._create_sheet("Config")

        current_row = 1

        # Title
        self._write_cell(current_row, 1, "System Configuration & Metadata", "header")
        current_row += 2

        # Configuration data
        self._write_cell(current_row, 1, "Run Information", "subheader")
        current_row += 1

        config_data = [
            ["Report Date", as_of.strftime("%Y-%m-%d")],
            ["MSA", msa],
            ["Asset", asset],
            ["Model Version", "2.1.0"],
            ["Data Vintage", "2024-Q3"],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]

        for config in config_data:
            self._write_cell(current_row, 1, config[0])
            self._write_cell(current_row, 2, config[1])
            current_row += 1

        current_row += 1

        # Feature flags
        self._write_cell(current_row, 1, "Feature Flags", "subheader")
        current_row += 1

        features = [
            ["State Rule Packs", "Enabled"],
            ["Market Scoring", "Enabled"],
            ["Asset Fit Engine", "Enabled"],
            ["Risk Engine", "Enabled"]
        ]

        for feature in features:
            self._write_cell(current_row, 1, feature[0])
            self._write_cell(current_row, 2, feature[1])
            current_row += 1

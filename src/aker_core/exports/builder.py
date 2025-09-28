from __future__ import annotations

import os
import tempfile
from datetime import date
from pathlib import Path
from typing import Optional
from sqlalchemy.orm import Session

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .sheets import (
    OverviewSheet,
    MarketScorecardSheet,
    AssetFitSheet,
    DealArchetypesSheet,
    RiskSheet,
    OpsKPIsSheet,
    COUTIDPatternsSheet,
    ChecklistSheet,
    DataLineageSheet,
    ConfigSheet
)


class ExcelWorkbookBuilder:
    """Main Excel workbook builder for Aker property model exports."""

    def __init__(self, session: Session):
        self.session = session
        self.workbook: Optional[Workbook] = None
        self.output_path: Optional[Path] = None

    def build_workbook(self, msa: str, asset: str, as_of: date) -> Path:
        """Build complete Excel workbook with all sheets.

        Args:
            msa: MSA code (e.g., "BOI", "DEN")
            asset: Asset identifier
            as_of: Date for data extraction

        Returns:
            Path to generated Excel file

        Raises:
            ValueError: If required data is missing or invalid
        """
        self.workbook = openpyxl.Workbook()
        self._setup_workbook_metadata(msa, asset, as_of)

        # Build each sheet
        self._build_overview_sheet(msa, asset, as_of)
        self._build_market_scorecard_sheet(msa, asset, as_of)
        self._build_asset_fit_sheet(msa, asset, as_of)
        self._build_deal_archetypes_sheet(msa, asset, as_of)
        self._build_risk_sheet(msa, asset, as_of)
        self._build_ops_kpis_sheet(msa, asset, as_of)
        self._build_coutid_patterns_sheet(msa, asset, as_of)
        self._build_checklist_sheet(msa, asset, as_of)
        self._build_data_lineage_sheet(msa, asset, as_of)
        self._build_config_sheet(msa, asset, as_of)

        # Add data connections
        self._add_database_connections()

        # Save workbook
        self.output_path = self._generate_output_path(msa, asset, as_of)
        self.workbook.save(self.output_path)
        return self.output_path

    def _setup_workbook_metadata(self, msa: str, asset: str, as_of: date):
        """Setup workbook metadata and properties."""
        if self.workbook is None:
            raise ValueError("Workbook not initialized")

        # Set workbook properties
        self.workbook.properties.title = f"Aker Property Model - {msa} - {asset}"
        self.workbook.properties.subject = f"Property analysis for {msa} as of {as_of}"
        self.workbook.properties.creator = "Aker Property Model"
        self.workbook.properties.created = as_of

        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

    def _build_overview_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Overview sheet."""
        sheet_builder = OverviewSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_market_scorecard_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Market_Scorecard sheet."""
        sheet_builder = MarketScorecardSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_asset_fit_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Asset_Fit sheet."""
        sheet_builder = AssetFitSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_deal_archetypes_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Deal_Archetypes sheet."""
        sheet_builder = DealArchetypesSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_risk_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Risk sheet."""
        sheet_builder = RiskSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_ops_kpis_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Ops_KPIs sheet."""
        sheet_builder = OpsKPIsSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_coutid_patterns_sheet(self, msa: str, asset: str, as_of: date):
        """Build the CO-UT-ID_Patterns sheet."""
        sheet_builder = COUTIDPatternsSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_checklist_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Checklist sheet."""
        sheet_builder = ChecklistSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_data_lineage_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Data_Lineage sheet."""
        sheet_builder = DataLineageSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _build_config_sheet(self, msa: str, asset: str, as_of: date):
        """Build the Config sheet."""
        sheet_builder = ConfigSheet(self.session, self.workbook)
        sheet_builder.build_sheet(msa, asset, as_of)

    def _add_database_connections(self):
        """Add database connections for live data."""
        if self.workbook is None:
            raise ValueError("Workbook not initialized")

        # Note: In a real implementation, this would add proper database connections
        # For now, we'll add placeholder connection information
        # In Excel, this would be done through external data connections
        pass

    def _generate_output_path(self, msa: str, asset: str, as_of: date) -> Path:
        """Generate output file path for the workbook."""
        # Create output directory if it doesn't exist
        output_dir = Path("exports")
        output_dir.mkdir(exist_ok=True)

        # Generate filename
        filename = f"aker_property_model_{msa}_{asset}_{as_of.strftime('%Y%m%d')}.xlsx"
        return output_dir / filename


# Convenience function for direct use
def to_excel(msa: str, asset: str, as_of: date, session: Optional[Session] = None) -> Path:
    """Generate Excel workbook for given parameters.

    Args:
        msa: MSA code (e.g., "BOI", "DEN")
        asset: Asset identifier
        as_of: Date for data extraction
        session: Database session (uses default if None)

    Returns:
        Path to generated Excel file

    Raises:
        ValueError: If required data is missing or invalid
    """
    if session is None:
        from aker_core.database import get_session
        session = get_session()

    builder = ExcelWorkbookBuilder(session)
    return builder.build_workbook(msa, asset, as_of)
